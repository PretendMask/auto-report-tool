import os
import sys
import glob
import tempfile
import traceback
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton,
    QVBoxLayout, QFileDialog, QProgressBar, QMessageBox
)
from PyQt6.QtCore import QThread, pyqtSignal, QObject
from PyQt6.QtGui import QFont, QIcon
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
import fitz  # PyMuPDF

# -------------------- 资源路径 --------------------
def resource_path(relative_path):
    """PyInstaller 单文件模式下获取资源路径"""
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# -------------------- PDF 渲染函数 --------------------
def pdf_to_image_pymupdf(pdf_path, output_path, zoom=2.0):
    try:
        doc = fitz.open(pdf_path)
        page = doc.load_page(0)
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        pix.save(output_path)
        doc.close()
        return True, None
    except Exception as e:
        return False, str(e)

# -------------------- Excel 工具函数 --------------------
def insert_image(ws, img_path, cell, row_idx, max_width=2.05*96, max_height=1.14*96):
    try:
        img = Image(img_path)
        width_ratio = max_width / img.width
        height_ratio = max_height / img.height
        scale = min(width_ratio, height_ratio)
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, cell)
        ws.row_dimensions[row_idx].height = img.height * 0.75
    except Exception as e:
        print(f"插入图片失败: {img_path}, 原因: {e}")

def set_cell_style(cell):
    cell.font = Font(name='宋体', size=16)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def extract_key(filename):
    name = os.path.basename(filename).rsplit('.', 1)[0]
    parts = name.split("-")
    if len(parts) < 3:
        return name
    return "-".join(parts[:3])

def find_order_file(files):
    for f in files:
        if "订单" in f and f.lower().endswith(".jpg"):
            return f
    return None

def find_invoice_file(files, exclude_files=None):
    exclude_files = exclude_files or []
    for f in files:
        if f in exclude_files:
            continue
        if "发票" in f and (f.lower().endswith(".jpg") or f.lower().endswith(".pdf")):
            return f
    return None

# -------------------- 自定义信号 --------------------
class WorkerSignals(QObject):
    progress_signal = pyqtSignal(int)
    file_signal = pyqtSignal(str)
    finished_signal = pyqtSignal(str)
    error_signal = pyqtSignal(str)

# -------------------- 后台线程处理 --------------------
class ProcessThread(QThread):
    def __init__(self, folder_path):
        super().__init__()
        self.folder_path = folder_path
        self.signals = WorkerSignals()

    def run(self):
        try:
            folder_path = self.folder_path
            excel_path = os.path.join(folder_path, "报销整理.xlsx")
            all_files = glob.glob(os.path.join(folder_path, "*.jpg")) + glob.glob(os.path.join(folder_path, "*.pdf"))
            if not all_files:
                self.signals.error_signal.emit("文件夹中未找到 JPG 或 PDF 文件")
                return

            file_dict = {}
            for f in all_files:
                key = extract_key(f)
                file_dict.setdefault(key, []).append(f)

            keys_list = list(file_dict.keys())
            total_files = len(keys_list)
            temp_files = []

            wb = Workbook()
            ws = wb.active
            ws.title = "报销"
            ws_headers = ["日期", "事由", "价格", "订单图", "发票"]
            ws.append(ws_headers)

            header_font = Font(name='宋体', size=16, bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for idx, cell in enumerate(ws[1], 1):
                cell.font = header_font
                cell.alignment = header_alignment
                if idx == 2:
                    ws.column_dimensions[cell.column_letter].width = 30
                elif idx == 3:
                    ws.column_dimensions[cell.column_letter].width = 12
                elif idx in [4,5]:
                    ws.column_dimensions[cell.column_letter].width = 2.05*96/7

            for current_index, key in enumerate(keys_list, start=1):
                files = file_dict[key]
                parts = key.split("-")
                date = parts[0] if len(parts) > 0 else ""
                reason = parts[1] if len(parts) > 1 else ""
                price = parts[2] if len(parts) > 2 else ""

                ws.append([date, reason, price])
                row_idx = ws.max_row
                for col_idx in range(1,6):
                    set_cell_style(ws.cell(row=row_idx, column=col_idx))

                # 处理订单图
                order_file = find_order_file(files)
                order_file_path = None
                if order_file:
                    order_file_path = os.path.abspath(order_file)
                    if os.path.exists(order_file_path):
                        insert_image(ws, order_file_path, f"D{row_idx}", row_idx)

                # 处理发票图
                invoice_file = find_invoice_file(files, exclude_files=[order_file] if order_file else [])
                if invoice_file:
                    if invoice_file.lower().endswith(".pdf"):
                        temp_path = os.path.join(tempfile.gettempdir(), f"{os.path.basename(invoice_file)}.jpg")
                        success, error = pdf_to_image_pymupdf(invoice_file, temp_path)
                        if success and os.path.exists(temp_path):
                            temp_files.append(temp_path)
                            insert_image(ws, temp_path, f"E{row_idx}", row_idx)
                    else:
                        invoice_file_path = os.path.abspath(invoice_file)
                        if os.path.exists(invoice_file_path):
                            insert_image(ws, invoice_file_path, f"E{row_idx}", row_idx)

                # 发送当前处理文件信号
                self.signals.file_signal.emit(key)
                progress_percent = int(current_index / total_files * 100)
                self.signals.progress_signal.emit(progress_percent)

            # 合计
            price_col_idx = 3
            sum_row_idx = ws.max_row + 1
            ws.cell(row=sum_row_idx, column=2, value="合计")
            total = 0.0
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=price_col_idx, max_col=price_col_idx):
                val = row[0].value
                try:
                    if val is not None and val != "":
                        total += float(val)
                except:
                    pass
            ws.cell(row=sum_row_idx, column=price_col_idx, value=total)
            for col_idx in range(1,6):
                cell = ws.cell(row=sum_row_idx, column=col_idx)
                set_cell_style(cell)
                if col_idx == price_col_idx:
                    cell.font = Font(name='宋体', size=16, bold=True)

            wb.save(excel_path)

            # 删除临时图片
            for temp_path in temp_files:
                if os.path.exists(temp_path):
                    os.remove(temp_path)

            self.signals.finished_signal.emit(excel_path)

        except Exception as e:
            self.signals.error_signal.emit(str(e))
            traceback.print_exc()

# -------------------- GUI --------------------
class ExpenseApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("简单报销整理")
        self.resize(600, 250)

        # 设置窗口图标
        self.setWindowIcon(QIcon(resource_path("icon.ico")))

        self.folder_path = ""

        self.layout = QVBoxLayout()
        self.label = QLabel("选择要处理的文件夹")
        self.label.setFont(QFont("Arial", 14))
        self.layout.addWidget(self.label)

        self.select_btn = QPushButton("选择文件夹")
        self.select_btn.clicked.connect(self.select_folder)
        self.layout.addWidget(self.select_btn)

        self.current_file_label = QLabel("当前处理文件：")
        self.layout.addWidget(self.current_file_label)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.layout.addWidget(self.progress)

        self.start_btn = QPushButton("开始处理")
        self.start_btn.clicked.connect(self.start_processing)
        self.layout.addWidget(self.start_btn)

        self.setLayout(self.layout)

    def select_folder(self):
        path = QFileDialog.getExistingDirectory(self, "选择文件夹", os.getcwd())
        if path:
            self.folder_path = path
            self.label.setText(f"当前文件夹: {path}")

    def start_processing(self):
        if not self.folder_path:
            QMessageBox.warning(self, "提示", "请先选择文件夹")
            return

        self.start_btn.setEnabled(False)
        self.thread = ProcessThread(self.folder_path)
        self.thread.signals.progress_signal.connect(self.update_progress)
        self.thread.signals.file_signal.connect(self.update_current_file)
        self.thread.signals.finished_signal.connect(self.processing_finished)
        self.thread.signals.error_signal.connect(self.processing_error)
        self.thread.start()

    def update_progress(self, value):
        self.progress.setValue(value)

    def update_current_file(self, filename):
        self.current_file_label.setText(f"当前处理文件：{filename}")

    def processing_finished(self, excel_path):
        self.start_btn.setEnabled(True)
        self.progress.setValue(100)
        self.current_file_label.setText("处理完成！")
        QMessageBox.information(self, "完成", f"报销整理完成，已保存到：\n{excel_path}")

    def processing_error(self, msg):
        self.start_btn.setEnabled(True)
        QMessageBox.critical(self, "错误", f"处理失败: {msg}")

# -------------------- 启动 --------------------
if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExpenseApp()
    window.show()
    sys.exit(app.exec())
